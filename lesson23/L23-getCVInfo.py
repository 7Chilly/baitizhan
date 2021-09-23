import os
import docx
import openpyxl

path = "teacher"
fileNames = os.listdir(path)
wb = openpyxl.load_workbook("汇总.xlsx")
aSheet = wb["教师信息登记表"]

count = 2
for item in fileNames:
    extension = os.path.splitext(item)[1]
    if extension != ".docx":
        continue
    filePath = os.path.join(path, item)
    teacherFile = docx.Document(filePath)
    table = teacherFile.tables[0]
    name = table.cell(0, 1).text
    gender = table.cell(0, 3).text
    birthDate = table.cell(0, 5).text
    highSchool = table.cell(3, 1).text
    masterSchool = table.cell(4, 1).text
    doctorSchool = table.cell(5, 1).text
    bestGraduate = table.cell(5, 5).text
    teachTime = table.cell(2, 1).text
    phoneNumber = table.cell(6, 1).text
    email = table.cell(6, 5).text
    job = table.cell(2, 5).text

    aSheet[f"A{count}"].value = name
    aSheet[f"B{count}"].value = gender
    aSheet[f"C{count}"].value = birthDate
    aSheet[f"D{count}"].value = highSchool
    aSheet[f"E{count}"].value = masterSchool
    aSheet[f"F{count}"].value = doctorSchool
    aSheet[f"G{count}"].value = bestGraduate
    aSheet[f"H{count}"].value = teachTime
    aSheet[f"I{count}"].value = job
    aSheet[f"J{count}"].value = phoneNumber
    aSheet[f"K{count}"].value = email

    print(f"{item}写入完成")

    count += 1

wb.save("汇总.xlsx")
