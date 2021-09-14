import os
import openpyxl
import docx

EnglishFilePath = "英语"
allItems = os.listdir(EnglishFilePath)
allStudentData = []
for item in allItems:
    docPath = os.path.join(EnglishFilePath, item)
    doc = docx.Document(docPath)
    table = doc.tables[0]
    studentData = {}
    studentData["name"] = table.cell(0,1).text
    studentData["gender"] = table.cell(0,3).text
    studentData["hometown"] = table.cell(0,5).text
    studentData["school"] = table.cell(1,1).text
    studentData["grade"] = table.cell(1,3).text
    studentData["major"] = table.cell(1,5).text
    studentData["tel"] = table.cell(2,1).text
    studentData["qq"] = table.cell(2,3).text
    studentData["yesorno"] = table.cell(2,5).text
    allStudentData.append(studentData)

wb = openpyxl.load_workbook("志愿者统计.xlsx")
sheet = wb["英语"]
for index, studentData in enumerate(allStudentData, 2):
    sheet[f"A{index}"].value = studentData["name"]
    sheet[f"B{index}"].value = studentData["gender"]
    sheet[f"C{index}"].value = studentData["hometown"]
    sheet[f"D{index}"].value = studentData["school"]
    sheet[f"E{index}"].value = studentData["grade"]
    sheet[f"F{index}"].value = studentData["major"]
    sheet[f"G{index}"].value = studentData["tel"]
    sheet[f"H{index}"].value = studentData["qq"]
    sheet[f"I{index}"].value = studentData["yesorno"]
wb.save("志愿者统计.xlsx")

