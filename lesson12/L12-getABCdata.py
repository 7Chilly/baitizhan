import openpyxl
from openpyxl import utils
newWb = openpyxl.Workbook()
aSheet = newWb["Sheet"]
aSheet.title = "A平台"
bSheet = newWb.create_sheet("B平台")
cSheet = newWb.create_sheet("C平台")
for sheet in newWb.worksheets:
    sheet["A1"].value = "商品名"
    sheet["B1"].value = "月份"
    sheet["C1"].value = "销售额"


def processMonthFile(filePath, orderSheetName, productIndex, orderHead, priceColumn, month, targetSheet):

    wb = openpyxl.load_workbook(filePath, data_only=True)
    orderSheet = wb[orderSheetName]
    MonthData = {}
    for rowData in orderSheet.rows:
        productName = rowData[productIndex].value
        if productName == orderHead:
            continue
        priceIndex = openpyxl.utils.cell.column_index_from_string(priceColumn) - 1
        price = rowData[priceIndex].value
        productGMV = MonthData.get(productName, 0)
        MonthData[productName] = productGMV + price

    for item in MonthData.keys():
        rowValue = [item, f"{month}月份", MonthData[item]]
        targetSheet.append(rowValue)
for month in range(1, 13):
    processMonthFile(f"A平台/2019{month:02}.xlsx","明细",4,"名称","F",month,aSheet)
    processMonthFile(f"B平台/order_2019_{month}.xlsx","订单详情",1,"商品名称","G",month,bSheet)
    processMonthFile(f"C平台/2019年{month}月销售订单.xlsx","销售订单数据",2,"商品名","I",month,cSheet)

newWb.save("汇总.xlsx")
