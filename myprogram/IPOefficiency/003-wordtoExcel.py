import openpyxl
import docx
from openpyxl import utils

wb = openpyxl.Workbook()
doc = docx.Document("【有研硅】财务部分.docx")
for idx, table in enumerate(doc.tables):
    if idx == 0:
        newSheet = wb["Sheet"]
        for rowIndex in range(1, len(table.rows)+1):
            for columnNumIndex in range(1, len(table.columns)+1):
                    columnIndex = utils.get_column_letter(columnNumIndex)
                    newSheet[f"{columnIndex}{rowIndex}"].value = table.cell(rowIndex-1, columnNumIndex-1).text
                    print(table.cell(rowIndex-1, columnNumIndex-1).text)
                    print(newSheet[f"{columnIndex}{rowIndex}"].value)
    else:
        if idx < 32 or 32 < idx < 49 or idx > 49:
            newSheet = wb.create_sheet(f"Sheet{idx}")
            for rowIndex in range(1, len(table.rows)+1):
                for columnNumIndex in range(1, len(table.columns)+1):
                    columnIndex = utils.get_column_letter(columnNumIndex)
                    newSheet[f"{columnIndex}{rowIndex}"].value = table.cell(rowIndex-1, columnNumIndex-1).text
                    print(table.cell(rowIndex-1, columnNumIndex-1).text)
                    print(newSheet[f"{columnIndex}{rowIndex}"].value)
        if idx == 32 and idx == 49:
            newSheet = wb.create_sheet(f"Sheet{idx}")
            wb.save("尝试表格.xlsx")
    wb.save("尝试表格.xlsx")
wb.save("尝试表格.xlsx")

# 目前没解决的问题有：
# word的表格写入excel后，字体等格式不知道如何保留。
# 遇到合并单元格的时候，word中的表格会输出重复单元格。
