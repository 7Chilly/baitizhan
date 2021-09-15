import openpyxl
import docx
from openpyxl import utils

wb = openpyxl.Workbook()
doc = docx.Document("九、主要财务指标.docx")
for idx, table in enumerate(doc.tables):
    if idx == 0:
        newSheet = wb["Sheet"]
    else:
        newSheet = wb.create_sheet()
    for rowIndex in range(1, len(table.rows)+1):
        for columnNumIndex in range(1, len(table.columns)+1):
            columnIndex = utils.get_column_letter(columnNumIndex + 1)
            newSheet[f"{columnIndex}{rowIndex}"].value = table.cell(rowIndex-1, columnNumIndex-1).text
wb.save("第一次尝试.xlsx")

# 目前没解决的问题有：
# word的表格写入excel后，字体等格式不知道如何保留。
# 遇到合并单元格的时候，word中的表格会输出重复单元格。
