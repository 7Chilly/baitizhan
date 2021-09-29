import openpyxl
import os

def getRowValues(self,row):
        columns = self.ws.max_column
        rowdata=[]
        for i in range(1,columns+1):
            cellvalue = self.ws.cell(row=row,column=i).value
            rowdata.append(cellvalue)
        return rowdata

wbpath = "cannabis空间测算.xlsx"
name = os.path.splitext(wbpath)[0]
wb = openpyxl.load_workbook(wbpath, data_only=True)
ws = wb["加拿大市场情况"]
maxColumnIndex = openpyxl.utils.cell.column_index_from_string("EQ") - 1

illegalList = ["非法市场"]
for i in range(2,maxColumnIndex+1,4):
    Q1 = ws.cell(6, i).value
    Q2 = ws.cell(6, i+1).value
    Q3 = ws.cell(6, i+2).value
    Q4 = ws.cell(6, i+3).value
    yearlyIllegal = Q1 + Q2 + Q3 + Q4
    illegalList.append(yearlyIllegal)

ws.append(illegalList)

newpath = name + "-改.xlsx"
wb.save(newpath)

