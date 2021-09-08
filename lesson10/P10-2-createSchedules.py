import openpyxl
wb = openpyxl.Workbook()
aSheet = wb["Sheet"]
aSheet.title = "七年级1班课表"
for i in range(2, 11):
    sheetName = f"七年级{i}班课表"
    wb.create_sheet(sheetName)

for sheet in wb.worksheets:
    sheet.merge_cells("A2:A4")
    sheet["A2"].value = "上午"
    sheet.merge_cells("A5:A6")
    sheet["A5"].value = "下午"
    sheet["B1"].value = "星期一"
    sheet["C1"].value = "星期二"
    sheet["D1"].value = "星期三"
    sheet["E1"].value = "星期四"
    sheet["F1"].value = "星期五"

wb.save("七年级课表.xlsx")
